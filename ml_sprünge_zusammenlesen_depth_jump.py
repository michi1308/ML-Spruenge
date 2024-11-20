import os
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.scrolledtext import ScrolledText
import pandas as pd
from openpyxl import load_workbook

# Hilfsfunktion zur Ausgabe in das Text-Widget
def output_to_widget(text_widget, message):
    text_widget.insert(tk.END, message + "\n")
    text_widget.see(tk.END)
    text_widget.update_idletasks()  # Aktualisiert das Textfeld sofort


def daten_zusammenlesen(folder_path, text_widget):
    # Liste aller Excel-Dateien im Ordner, die nicht temporär sind
    excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx') and not f.startswith('~$')]

    # Leere Liste, um alle transponierten Tabellen zu speichern
    all_transposed_data = []

    # Verarbeitung der einzelnen Excel-Dateien
    for file in excel_files:
        file_path = os.path.join(folder_path, file)

        try:
            # Lade die Excel-Datei in einen Pandas DataFrame (erste Tabelle wird standardmäßig geladen)
            df = pd.read_excel(file_path, header=None)

            # Extrahiere die Werte für die neuen Spalten
            code_value = df.iloc[8, 1]  # B9
            nachname_value = df.iloc[1, 1]  # B2
            vorname_value = df.iloc[2, 1]  # B3
            geschlecht_value = df.iloc[3, 1]  # B4
            groeße_value = df.iloc[6, 1]  # B7
            gewicht_value = df.iloc[5, 1]  # B6
            dominantes_bein_value = df.iloc[7, 1]  # B8

            # Ermittlung der letzten Spalte der Datei
            last_col_index = df.shape[1] - 1  # Vorletzte Spalte, weil der Index bei 0 beginnt

            # Extrahiere die Daten aus den gewünschten Bereichen:
            text_values = df.iloc[18:68, 0].tolist()  # A19 bis A68
            num_values_1 = df.iloc[18:68, 1].tolist()  # B19 bis B68

            # Dynamische Extraktion der weiteren Werte bis zur vorletzten Spalte
            all_num_values = [df.iloc[18:68, i].tolist() for i in range(2, last_col_index)]  # Ab Spalte C bis zur vorletzten Spalte

            # Erstelle das DataFrame für die extrahierten Daten
            data_dict = {'Text': text_values, 'Wert1': num_values_1}
            # Füge die weiteren Werte dynamisch hinzu
            for i, values in enumerate(all_num_values, 2):  # ab Spalte 2 (Wert2, Wert3, ...)
                data_dict[f'Wert{i}'] = values

            # Erstelle DataFrame aus dem Dictionary
            data_df = pd.DataFrame(data_dict)

            # Transponiere das DataFrame
            transposed_data = data_df.T

            # Füge die Beschriftungen als erste Zeile hinzu
            transposed_df = pd.DataFrame(transposed_data.values, columns=transposed_data.iloc[0])
            transposed_df = transposed_df[1:]  # Entferne die erste Zeile, die die Beschriftungen enthält

            # Bereite die Werte für die neuen Spalten vor
            num_rows = transposed_df.shape[0]

            # Leere Felder für die neuen Variablen mit Werten aus den spezifischen Zellen
            code = [code_value for _ in range(num_rows)]
            vorname = [vorname_value for _ in range(num_rows)]
            nachname = [nachname_value for _ in range(num_rows)]
            geschlecht = [geschlecht_value for _ in range(num_rows)]
            groeße = [groeße_value for _ in range(num_rows)]
            gewicht = [gewicht_value for _ in range(num_rows)]
            dominantes_bein = [dominantes_bein_value for _ in range(num_rows)]

            if num_rows > 0:
                # Erstelle das DataFrame mit den zusätzlichen Spalten
                additional_columns = pd.DataFrame({
                    'Code': code,
                    'Vorname': vorname,
                    'Nachname': nachname,
                    'Geschlecht': geschlecht,
                    'Größe': groeße,
                    'Gewicht': gewicht,
                    'Dominantes Bein': dominantes_bein
                })

                # Füge die zusätzlichen Spalten an den Anfang der transponierten Tabelle hinzu
                final_df = pd.concat([additional_columns, transposed_df.reset_index(drop=True)], axis=1)

                # Speichere das transponierte DataFrame in der Liste
                all_transposed_data.append(final_df)
            else:
                output_to_widget(text_widget, f"Die Datei {file} hat keine verarbeitbaren Daten.")

        except Exception as e:
            output_to_widget(text_widget, f"Fehler beim Verarbeiten der Datei {file}: {e}")

    # Überprüfe, ob es Daten gibt, bevor du versuchst zu concatenieren
    if all_transposed_data:
        # Fasse alle transponierten Tabellen in eine einzige Tabelle zusammen
        final_data = pd.concat(all_transposed_data, ignore_index=True)

        # Speichere die finale Tabelle als neue Excel-Datei
        output_file_path = os.path.join(folder_path, 'ML_zusammengeführte_Datei.xlsx')
        final_data.to_excel(output_file_path, index=False)

        # Lade die erstellte Datei mit openpyxl
        wb = load_workbook(output_file_path)
        ws = wb.active  # Aktiviere das Arbeitsblatt

        # Lösche die Spalten AS, AT und AU
        ws.delete_cols(ws["AU1"].col_idx)  # Lösche Spalte AU
        ws.delete_cols(ws["AT1"].col_idx)  # Lösche Spalte AT
        ws.delete_cols(ws["AS1"].col_idx)  # Lösche Spalte AS

        # Setze die Breite für alle Spalten
        for col in ws.columns:
            col_letter = col[0].column_letter  # Erhalte den Spaltenbuchstaben
            ws.column_dimensions[col_letter].width = 40  # Setze die Spaltenbreite auf 40

        # Speichere die Excel-Datei mit den aktualisierten Spaltenbreiten
        wb.save(output_file_path)
        wb.close()  # Schließe die Arbeitsmappe

        output_to_widget(text_widget, f"Alle Dateien wurden erfolgreich verarbeitet und die Datei wurde gespeichert.")
        messagebox.showinfo("Erfolg", f"Die Tabelle wurde erfolgreich erstellt und gespeichert unter: {folder_path}")


# GUI-Setup
def ordner_auswaehlen(entry):
    folder_selected = filedialog.askdirectory()
    entry.delete(0, tk.END)
    entry.insert(0, folder_selected)

def verabeitung_starten(entry, text_widget):
    folder_path = entry.get()
    if not os.path.isdir(folder_path):
        messagebox.showerror("Fehler", "Bitte geb einen gültigen Ordnerpfad an.")
        return
    text_widget.delete(1.0, tk.END)  # Löscht die Textausgabe
    daten_zusammenlesen(folder_path, text_widget)

def main():
    # Hauptfenster erstellen
    root = tk.Tk()
    root.title("ML Sprungdateien zusammenlesen")

    # Eingabefeld für Ordnerpfad
    frame = tk.Frame(root)
    frame.pack(padx=10, pady=10)

    entry_label = tk.Label(frame, text="Bitte den Ordnerpfad angeben:")
    entry_label.grid(row=0, column=0, sticky="w")

    # Einstellen der Breite des Eingabefeldes
    entry = tk.Entry(frame, width=70)
    entry.grid(row=0, column=1)

    browse_button = tk.Button(frame, text="Durchsuchen", command=lambda: ordner_auswaehlen(entry))
    browse_button.grid(row=0, column=2, padx=5)

    # Start-Button
    start_button = tk.Button(frame, text="Starten", command=lambda: verabeitung_starten(entry, text_output))
    start_button.grid(row=1, column=1, pady=10)

    # Text-Widget für die Ausgaben
    text_output = ScrolledText(root, height=20, width=100)
    text_output.pack(padx=10, pady=10)

    root.mainloop()

# Hauptprogramm starten
if __name__ == "__main__":
    main()
