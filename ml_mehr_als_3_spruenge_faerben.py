import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def lade_excel_datei(dateipfad):
    """
    Funktion, um eine Excel-Datei als Pandas DataFrame zu laden.
    """
    try:
        df = pd.read_excel(dateipfad, header=None)
        return df
    except Exception as e:
        raise Exception(f"Fehler beim Laden der Datei: {e}")


def finde_haeufige_werte(df, spalte_index, grenze=3):
    """
    Funktion, die häufige Werte in einer bestimmten Spalte identifiziert.
    Gibt eine Liste der Zeilenindizes zurück, in denen die Werte häufiger vorkommen als `grenze`.
    """
    value_counts = df.iloc[:, spalte_index].value_counts()
    haeufige_werte = value_counts[value_counts > grenze].index
    zeilen_zu_faerben = df[df.iloc[:, spalte_index].isin(haeufige_werte)].index.tolist()
    return zeilen_zu_faerben


def faerbe_zellen(dateipfad, zeilen_zu_faerben, spalte_index, farbe="FFFF00"):
    """
    Funktion, die bestimmte Zellen in einer Excel-Datei mit einer Farbe markiert.
    """
    try:
        wb = load_workbook(dateipfad)
        ws = wb.active

        # Gelbe Füllung definieren
        fill = PatternFill(start_color=farbe, end_color=farbe, fill_type="solid")

        # Zellen in der angegebenen Spalte färben
        for zeile in zeilen_zu_faerben:
            zellenadresse = ws.cell(row=zeile + 1, column=spalte_index + 1)  # Excel ist 1-basiert
            zellenadresse.fill = fill

        # Speichere die Datei
        wb.save(dateipfad)
        wb.close()
    except Exception as e:
        raise Exception(f"Fehler beim Bearbeiten der Datei: {e}")


def main():
    dateipfad = r"K:\Team\Böhmer_Michael\Tabellen_ML_Sprünge_Patient\ML_CMJ.xlsx"
    spalte_index = 0  # Erste Spalte
    grenze = 3  # Werte, die öfter als dreimal vorkommen

    try:
        # 1. Lade die Datei
        df = lade_excel_datei(dateipfad)

        # 2. Finde die Zeilen, die gefärbt werden müssen
        zeilen_zu_faerben = finde_haeufige_werte(df, spalte_index, grenze)

        # 3. Färbe die Zellen
        faerbe_zellen(dateipfad, zeilen_zu_faerben, spalte_index)

        print(f"Die Datei wurde erfolgreich bearbeitet. {len(zeilen_zu_faerben)} Zellen wurden gelb gefärbt.")

    except Exception as e:
        print(f"Fehler: {e}")


if __name__ == "__main__":
    main()
